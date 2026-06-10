"""
make_tables.py — regenerates Tables 1-3 as .xlsx and .tex (booktabs, ready to \\input).
Table 3 values come from the Langevin ensembles in make_figures.py (n=5000, seed 42,
splitter at y_s = -39.7 um).

Usage:
    python make_tables.py [outdir]   (default: current directory)

Outputs:
    table1_design_parameters.{xlsx,tex}
    table2_force_budget.{xlsx,tex}
    table3_separation_performance.{xlsx,tex}

Requirements: openpyxl
"""

import os, sys, textwrap
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Shared Excel helpers
HEADER_FILL  = PatternFill('solid', fgColor='D9D9D9')
SECTION_FILL = PatternFill('solid', fgColor='F2F2F2')

def _hdr(ws, row, cols, values):
    for c, val in zip(cols, values):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font      = Font(bold=True, name='Arial', size=10)
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def _sec(ws, row, col_span, label):
    cell = ws.cell(row=row, column=1, value=label)
    cell.font      = Font(bold=True, name='Arial', size=9, italic=True)
    cell.fill      = SECTION_FILL
    if col_span > 1:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=col_span)
    cell.alignment = Alignment(horizontal='left')

def _row(ws, row, cols, values, bold=False):
    for c, val in zip(cols, values):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font      = Font(bold=bold, name='Arial', size=9)
        cell.alignment = Alignment(horizontal='left', wrap_text=True)

def _autowidth(ws, min_w=10, max_w=50):
    for col in ws.columns:
        length = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(length * 1.15, min_w), max_w)

def _save_xlsx(wb, outdir, name):
    os.makedirs(outdir, exist_ok=True)
    path = os.path.join(outdir, name)
    wb.save(path)
    print(f'  Saved  {name}')

def _save_tex(tex_str, outdir, name):
    os.makedirs(outdir, exist_ok=True)
    path = os.path.join(outdir, name)
    with open(path, 'w', encoding='utf-8') as fh:
        fh.write(tex_str)
    print(f'  Saved  {name}')


# Table 1 — Design Parameters (grouped by category; Ito-corrected pressure drop only)

T1_GROUPS = [
    ('Channel geometry', [
        ('Channel depth (axial)',          r'$w$',          r'500\,\textmu m'),
        ('Channel gap (radial, migration axis)', r'$h$',  r'200\,\textmu m'),
        ('Aspect ratio',                   r'$h/w$',        r'0.40'),
        ('Hydraulic diameter',             r'$D_h$',        r'285.7\,\textmu m'),
    ]),
    ('Spiral path', [
        ('Inner radius',                   r'$R_\text{in}$',  r'3.0\,mm'),
        ('Outer radius',                   r'$R_\text{out}$', r'7.5\,mm'),
        ('Number of turns',                r'$N$',            r'5'),
        ('Path length',                    r'$L$',            r'164.9\,mm'),
    ]),
    ('Operating point', [
        ('Mean axial velocity',            r'$U$',            r'0.55\,m\,s$^{-1}$'),
        ('Volumetric flow rate',           r'$Q$',            r'3.30\,mL\,min$^{-1}$'),
        ('Residence time',                 r'$\tau$',         r'0.300\,s'),
        ('Channel Reynolds number',        r'$\mathrm{Re}_c$',r'151'),
        ('Particle Reynolds number',       r'$\mathrm{Re}_p$',r'0.118'),
        ('Dean number (outer to inner)',   r'$\mathrm{De}$',  r'20.8\,--\,32.8'),
        ('Pressure drop (curvature-corrected)', r'$\Delta P$', r'43.9\,kPa'),
    ]),
    ('Particles', [
        ('Diameter (both species)',        r'$a$',            r'8\,\textmu m'),
        ('Confinement ratio',              r'$a/D_h$',        r'0.028'),
        ('PTFE density',                   r'$\rho_\text{PTFE}$', r'2200\,kg\,m$^{-3}$'),
        ('PVDF density',                   r'$\rho_\text{PVDF}$', r'1780\,kg\,m$^{-3}$'),
    ]),
    (r'Carrier fluid  (3.5\% NaCl, 25\,\textdegree C)', [
        ('Fluid density',                  r'$\rho_f$',   r'1025\,kg\,m$^{-3}$'),
        ('Dynamic viscosity',              r'$\eta$',     r'$1.070\times10^{-3}$\,Pa\,s'),
    ]),
]


def table1_design_parameters(outdir):
    # --- Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Design Parameters'
    ws.sheet_view.showGridLines = False
    _hdr(ws, 1, [1, 2, 3], ['Quantity', 'Symbol', 'Value'])

    r = 2
    for group_name, rows in T1_GROUPS:
        _sec(ws, r, 3, group_name.replace(r'\textdegree ', '°')
                                 .replace(r'\%', '%')
                                 .replace(r'\,', ' '))
        r += 1
        for qty, sym, val in rows:
            _row(ws, r, [1, 2, 3], [
                qty,
                sym.replace('$', '').replace('\\', '').replace('_', '').replace('{', '').replace('}', ''),
                val.replace('$', '').replace('\\', '').replace(',', ' ').replace('{', '').replace('}', ''),
            ])
            r += 1

    _autowidth(ws)
    _save_xlsx(wb, outdir, 'table1_design_parameters.xlsx')

    # --- LaTeX ---
    lines = [
        r'\begin{table}[htbp]',
        r'\centering',
        r'\caption{Design parameters at the operating point.}',
        r'\label{tab:design}',
        r'\begin{tabular}{@{}llr@{}}',
        r'\toprule',
        r'Quantity & Symbol & Value \\',
        r'\midrule',
    ]
    for group_name, rows in T1_GROUPS:
        lines.append(
            r'\multicolumn{3}{@{}l}{\textit{' + group_name + r'}} \\')
        for qty, sym, val in rows:
            lines.append(f'{qty} & {sym} & {val} \\\\')
        lines.append(r'\addlinespace[2pt]')

    # remove last \addlinespace before \bottomrule
    if lines[-1] == r'\addlinespace[2pt]':
        lines.pop()

    lines += [
        r'\bottomrule',
        r'\end{tabular}',
        r'\end{table}',
    ]
    _save_tex('\n'.join(lines) + '\n', outdir, 'table1_design_parameters.tex')


# Table 2 — Force Budget

def table2_force_budget(outdir):
    # --- Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Force Budget'
    ws.sheet_view.showGridLines = False

    _hdr(ws, 1, [1, 2, 3, 4, 5],
         ['Force', 'Scaling', 'PTFE (pN)', 'PVDF (pN)', 'In force balance?'])

    _sec(ws, 2, 5, 'Force-balance contributions')
    _row(ws, 3, [1, 2, 3, 4, 5],
         ['Inertial lift  F_L  (f_L = 0.5)',
          'proportional to a^4', '7.8', '7.8', 'Yes'])
    _row(ws, 4, [1, 2, 3, 4, 5],
         ['Centrifugal buoyancy  F_B',
          'proportional to delta-rho', '18.1', '11.7', 'Yes'])

    _sec(ws, 5, 5, 'Discriminating signal')
    _row(ws, 6, [1, 2, 3, 4, 5],
         ['Delta F_B  (PTFE minus PVDF)',
          '—', '6.5', '—', 'Sole species-dependent term'], bold=True)

    _sec(ws, 7, 5, 'Present but non-contributing  (a/D_h = 0.028 < 0.07 threshold)')
    _row(ws, 8, [1, 2, 3, 4, 5],
         ['Dean drag  F_D',
          'proportional to a', '2728', '2728',
          'No — particles cycle with Dean vortex; time-averaged drift = 0'])

    ws.row_dimensions[8].height = 30
    _autowidth(ws)
    _save_xlsx(wb, outdir, 'table2_force_budget.xlsx')

    # --- LaTeX ---
    tex = textwrap.dedent(r"""
    \begin{table}[htbp]
    \centering
    \caption{Force budget for a single 8\,\textmu m particle at the design point
             ($R_c = 5.25$\,mm).  Forces in piconewtons.}
    \label{tab:forces}
    \begin{tabular}{@{}llSSl@{}}
    \toprule
    Force & Scaling & {PTFE (pN)} & {PVDF (pN)} & In force balance? \\
    \midrule
    Inertial lift $F_L$\;($f_L=0.5$) & $\propto a^4$          &  7.8 &  7.8 & Yes \\
    Centrifugal buoyancy $F_B$       & $\propto \Delta\rho$   & 18.1 & 11.7 & Yes \\
    \midrule
    \textbf{Discriminating signal} $\bm{\Delta F_B}$ & --- & \textbf{6.5} & --- &
        \textit{sole species-dependent term} \\
    \midrule
    Dean drag $F_D$                  & $\propto a$            & 2728 & 2728 & No$^{a}$ \\
    \bottomrule
    \end{tabular}
    \begin{tablenotes}\footnotesize
    \item[$^{a}$] $a/D_h = 0.028 < 0.07$ (Di~Carlo 2007 threshold); particles
    cycle with the Dean vortex and their time-averaged lateral drift is zero.
    \end{tablenotes}
    \end{table}
    """).strip()
    _save_tex(tex + '\n', outdir, 'table2_force_budget.tex')


# Table 3 — Separation Performance

def table3_separation_performance(outdir):
    # --- Excel ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Separation Performance'
    ws.sheet_view.showGridLines = False

    _hdr(ws, 1, [1, 2, 3, 4],
         ['Outlet', 'Target species', 'Purity (%)', 'Recovery'])

    _sec(ws, 2, 4, 'Uniform inlet (control), y0 in [-85,+85] um')
    _row(ws, 3, [1, 2, 3, 4], ['Outer', 'PTFE', '62.9',  '1.000'])
    _row(ws, 4, [1, 2, 3, 4], ['Inner', 'PVDF', '100.0', '0.410'])

    _sec(ws, 5, 4, 'Focused inlet (design, sheath flow), y0 in [-85,-15] um')
    _row(ws, 6, [1, 2, 3, 4], ['Outer', 'PTFE', '98.5',  '1.000'])
    _row(ws, 7, [1, 2, 3, 4], ['Inner', 'PVDF', '100.0', '0.985'])

    _autowidth(ws)
    _save_xlsx(wb, outdir, 'table3_separation_performance.xlsx')

    # --- LaTeX ---
    tex = textwrap.dedent(r"""
    \begin{table}[htbp]
    \centering
    \caption{Predicted separation performance from the Langevin ensembles
             ($n=5000$ per species, seed 42), single radial splitter at
             $y_s = -39.7$\,\textmu m.  Purity = fraction of particles at an
             outlet that belong to the target species; recovery = fraction of
             that species collected at its intended outlet.}
    \label{tab:sep}
    \begin{threeparttable}
    \begin{tabular}{@{}llcc@{}}
    \toprule
    Outlet & Target species & Purity (\%) & Recovery \\
    \midrule
    \multicolumn{4}{@{}l}{\textit{Uniform inlet (control), $y_0 \in [-85,+85]$\,\textmu m}} \\
    Outer & PTFE &  62.9 & 1.000 \\
    Inner & PVDF & 100.0 & 0.410 \\
    \addlinespace[3pt]
    \multicolumn{4}{@{}l}{\textit{Focused inlet (design, sheath flow), $y_0 \in [-85,-15]$\,\textmu m}} \\
    Outer & PTFE &  98.5 & 1.000 \\
    Inner & PVDF & 100.0 & 0.985 \\
    \bottomrule
    \end{tabular}
    \begin{tablenotes}[flushleft]
    \footnotesize
    \item Under uniform seeding ($y_0 \in [-85,+85]$\,\textmu m), about half of each
    species converges to the outer-branch equilibria (separated by only 2.0\,\textmu m)
    on the far side of the splitter, contaminating the outer stream and stranding
    PVDF; sheath-flow seeding ($y_0 \in [-85,-15]$\,\textmu m) routes both species
    to the well-separated inner branch and removes both loss channels.
    \end{tablenotes}
    \end{threeparttable}
    \end{table}
    """).strip()
    _save_tex(tex + '\n', outdir, 'table3_separation_performance.tex')


if __name__ == '__main__':
    outdir = sys.argv[1] if len(sys.argv) > 1 else '.'
    print('Tables 1-3  (Excel + LaTeX)')
    table1_design_parameters(outdir)
    table2_force_budget(outdir)
    table3_separation_performance(outdir)
    print('All done.')
